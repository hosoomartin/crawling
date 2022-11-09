# 실행전 주의사항
# fpath의 data.xlsx 위치를 본인에 맞게 수정할 것

import requests
from bs4 import BeautifulSoup
import openpyxl
import os

fpath = r'D:\Personal Folder\topdata.xlsx'

# 파일 있음 load, 없음 생성
if os.path.isfile(fpath) :
    wb = openpyxl.load_workbook(fpath)
else :  
    wb = openpyxl.Workbook()
# wb = openpyxl.load_workbook()
ws = wb.active  # 현재 활성화된 시트 선택

# 타이틀
ws['A1'] = '종목명'
ws['B1'] = '현재가'
ws['C1'] = '차액'
ws['D1'] = '비율'

rows = 10
url = f"https://finance.naver.com"

"""
# Selenium을 이용위해 추가 import
from selenium import webdriver

# 브라우저 생성
browser = webdriver.Chrome('C:/chromedriver.exe')  #사전에 다운로드 해 둬야 함

# Selenium으로 먼저 html 가져옴
browser.get(url_name)
html = browser.page_source
"""

response = requests.get(url)
html = response.text
soup = BeautifulSoup(html, 'html.parser')

for row in range(1, rows+1) :
    name = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > th > a').text
    price = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > td:nth-child(2)').text
    delta = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > td:nth-child(3)').text
    delta_rate = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > td:nth-child(4)').text
    print(f'{name}, {price}, {delta}, {delta_rate}')
    ws[f'A{row+1}'] = name
    ws[f'B{row+1}'] = price
    ws[f'C{row+1}'] = delta
    ws[f'D{row+1}'] = delta_rate

wb.save(fpath)