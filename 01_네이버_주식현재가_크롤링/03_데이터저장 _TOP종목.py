# 실행전 주의사항
# fpath의 data.xlsx 위치를 본인에 맞게 수정할 것

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles.fonts import Font
import os

fpath = r'D:\Personal Folder\topdata.xlsx'

# 파일 있음 load, 없음 생성
if os.path.isfile(fpath) :
    wb = openpyxl.load_workbook(fpath)
else :  
    wb = openpyxl.Workbook()
# wb = openpyxl.load_workbook()
ws = wb.active  # 현재 활성화된 시트 선택

tops = [
    '거래상위',
    '상승',
    '하락',
    '시가총액상위',
]

heads = [
    '종목명',
    '현재가',
    '차액',
    '비율',
]

top_rows = 10
current_row = 1
url = f"https://finance.naver.com"

response = requests.get(url)
html = response.text
soup = BeautifulSoup(html, 'html.parser')


font_tops = Font(size=12, bold=True)
font_head = Font(bold=True)

for top in tops:
    ws[f'A{current_row}'] = top
    ws[f'A{current_row}'].font = font_tops
    current_row += 1

    # column 세팅
    current_column = 1
    for head in heads:
        ws.cell(row=current_row, column=current_column).value = head
        current_column += 1

    current_row += 1
    for row in range(1, top_rows+1) :
        name = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > th > a').text
        price = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > td:nth-child(2)').text
        delta = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > td:nth-child(3)').text
        delta_rate = soup.select_one(f'#_topItems1 > tr:nth-child({row}) > td:nth-child(4)').text
        print(f'{name}, {price}, {delta}, {delta_rate}')
        ws[f'A{current_row}'] = name
        ws[f'B{current_row}'] = price
        ws[f'C{current_row}'] = delta
        ws[f'D{current_row}'] = delta_rate

wb.save(fpath)